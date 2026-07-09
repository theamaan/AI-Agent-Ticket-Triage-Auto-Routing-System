"""
Generates a rich 50-ticket sample Excel file and saves it to input/.
Run: python scripts/generate_sample_tickets.py
"""
from __future__ import annotations

import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Ticket data ───────────────────────────────────────────────────────────────

TICKETS = [
    # ── Provided tickets 1-30 ────────────────────────────────────────────────
    (1,  "P2", "Open",        "Initial 999 rejection during claim intake",
         "EDI mapping issue",
         "Claim processing error requiring correction. Functional acknowledgement returned rejection code for initial 999 transaction. Manual intervention needed to resubmit."),
    (2,  "P2", "In Progress", "Final 999 error in outbound batch",
         "EDI transaction error",
         "Batch execution failed impacting downstream systems. Final 999 file was not accepted by trading partner. Batch has been halted pending investigation."),
    (3,  "P3", "Open",        "File transfer failed to trading partner",
         "EDI partner connectivity",
         "Transaction failed due to missing mandatory segments. Secure file transfer to partner endpoint returned a connection timeout error after three retry attempts."),
    (4,  "P2", "Pending",     "Initial 999 rejection – invalid ISA segment",
         "EDI mapping issue",
         "Transaction failed due to missing mandatory segments. ISA header fields do not match the expected format defined in the trading partner agreement."),
    (5,  "P2", "Open",        "Claim rejected in 277CA response",
         "EDI transaction error",
         "Batch execution failed impacting downstream systems. 277CA file indicates claim-level rejection. Root cause identified as missing NPI in the 2010BB loop."),
    (6,  "P3", "In Progress", "Schema validation failure on inbound file",
         "HIPAA validation failure",
         "Transaction failed due to missing mandatory segments. Inbound 837P file failed XSD schema validation at the GS/GE envelope level."),
    (7,  "P1", "Open",        "BizTalk suspended message in EDI receive port",
         "BizTalk orchestration failure",
         "Issue observed in EDI pipeline requiring validation and reprocessing. Multiple messages suspended in BizTalk receive pipeline. Orchestration halted awaiting admin action."),
    (8,  "P2", "Pending",     "Claim rejected in 277CA – duplicate submission",
         "EDI transaction error",
         "File not processed due to format issue. 277CA shows claim rejected for duplicate submission. Original claim date and ID match a prior submission in the system."),
    (9,  "P1", "Open",        "BizTalk suspended message – schema mismatch",
         "Schema validation error",
         "Issue observed in EDI pipeline requiring validation and reprocessing. Message suspended due to schema resolution failure in the BizTalk pipeline stage."),
    (10, "P1", "In Progress", "BizTalk suspended message – orchestration abort",
         "BizTalk orchestration failure",
         "Acknowledgement failure detected and needs fix. Orchestration instance suspended in BizTalk after an unhandled exception during the 997 acknowledgement send shape."),
    (11, "P3", "Open",        "Data mismatch in X12 transaction set",
         "EDI mapping issue",
         "Issue observed in EDI pipeline requiring validation and reprocessing. X12 data element values do not match the expected codes defined in the partner configuration table."),
    (12, "P3", "Pending",     "Acknowledgement not received from trading partner",
         "Acknowledgement not received",
         "Acknowledgement failure detected and needs fix. 997 functional acknowledgement has not been received within the agreed SLA window of 2 hours."),
    (13, "P3", "Open",        "File transfer failed – SFTP credentials expired",
         "EDI partner connectivity",
         "Acknowledgement failure detected and needs fix. SFTP key authentication failed. Credentials must be rotated and re-registered with the trading partner."),
    (14, "P1", "In Progress", "EDI transaction failed during processing pipeline",
         "EDI transaction error",
         "Batch execution failed impacting downstream systems. Critical EDI processing job terminated mid-run. Database rollback initiated. Downstream claims processing is blocked."),
    (15, "P1", "Open",        "BizTalk suspended message – integration service down",
         "BizTalk host instance down",
         "Integration failure between systems. BizTalk host instance stopped unexpectedly. All in-flight messages suspended. Service restart required with message resume."),
    (16, "P2", "Pending",     "Claim rejected in 277CA – member not found",
         "EDI transaction error",
         "Reprocessing required after fixing data. 277CA indicates member ID does not match any active enrollment record. Eligibility check needed before resubmission."),
    (17, "P2", "Open",        "Batch job failure in Tidal scheduler",
         "EDI transaction error",
         "Integration failure between systems. Tidal batch job did not complete within the allotted window. Job log shows resource contention on the processing server."),
    (18, "P2", "In Progress", "EDI transaction failed during processing – invalid GS",
         "EDI mapping issue",
         "Reprocessing required after fixing data. GS segment functional identifier code does not match the transaction set type. File rejected at interchange level."),
    (19, "P1", "Open",        "Batch job failure in Tidal – critical downstream impact",
         "EDI transaction error",
         "Batch execution failed impacting downstream systems. Tidal job failure caused 48 downstream jobs to remain in queued state. Escalation required."),
    (20, "P2", "Pending",     "Claim rejected in 277CA – invalid provider",
         "EDI transaction error",
         "Reprocessing required after fixing data. Provider NPI in 277CA response is not enrolled in the payer system. Provider enrollment verification needed."),
    (21, "P3", "Open",        "Claim rejected in 277CA – format issue detected",
         "HIPAA validation failure",
         "File not processed due to format issue. 277CA response highlights a format violation in the 2200D loop. Date element is not in the required CCYYMMDD format."),
    (22, "P1", "In Progress", "BizTalk suspended message – 997 send failure",
         "BizTalk orchestration failure",
         "Acknowledgement failure detected and needs fix. 997 outbound acknowledgement message suspended in BizTalk send pipeline due to adapter configuration error."),
    (23, "P2", "Pending",     "BizTalk suspended message – reprocessing needed",
         "Suspended message in BizTalk",
         "Reprocessing required after fixing data. Three messages remain suspended after a prior failed reprocessing attempt. Root cause analysis in progress."),
    (24, "P2", "Open",        "EDI transaction failed during processing – timeout",
         "EDI transaction error",
         "Integration failure between systems. Database transaction timeout during EDI processing caused a partial write. Data consistency check required before retry."),
    (25, "P3", "Open",        "Acknowledgement not received – partner SLA breach",
         "Acknowledgement not received",
         "Issue observed in EDI pipeline requiring validation and reprocessing. Trading partner has exceeded the 997 response SLA. Escalation email sent to partner contact."),
    (26, "P2", "In Progress", "Initial 999 rejection – data validation failed",
         "EDI mapping issue",
         "Data validation failed and marked invalid. Initial 999 transaction rejected. Validation engine flagged the ISA13 control number as previously used. Deduplication required."),
    (27, "P1", "Open",        "EDI transaction failed during processing – memory error",
         "EDI transaction error",
         "Batch execution failed impacting downstream systems. Processing service threw an out-of-memory exception during large batch ingestion. Server resources need review."),
    (28, "P1", "In Progress", "EDI transaction failed – BizTalk orchestration abort",
         "BizTalk orchestration failure",
         "System generated error during BizTalk orchestration. Orchestration dehydration failed mid-processing. Instance ID captured for debugging. Suspend and resume required."),
    (29, "P3", "Pending",     "Schema validation failure – missing GE segment",
         "HIPAA validation failure",
         "File not processed due to format issue. GE segment is absent from the functional group footer. File structure is non-compliant with X12 standards."),
    (30, "P2", "Open",        "Final 999 error – acknowledgement failure",
         "EDI transaction error",
         "Acknowledgement failure detected and needs fix. Final 999 response not delivered to originating system. ACH notification pipeline also affected by the same connectivity gap."),
    # ── Extended tickets 31-50 ───────────────────────────────────────────────
    (31, "P3", "Open",        "SSIS package execution failed on nightly ETL",
         "EDI mapping issue",
         "SSIS ETL package failed at the data transformation step. Source table schema change broke the column mapping. Package must be updated and revalidated before next run."),
    (32, "P2", "In Progress", "MMW interface not responding to health check",
         "EDI partner connectivity",
         "MMW healthcare data exchange service is unreachable. Health check endpoint returning HTTP 503. Suspected network routing issue between application and integration servers."),
    (33, "P3", "Pending",     "SSRS report generation error – null reference",
         "EDI mapping issue",
         "SSRS report failed to render due to a null reference exception in the dataset query. Report parameter default value is resolving to NULL causing the stored procedure to fail."),
    (34, "P2", "Open",        "SSIS job aborted mid-run – connection pool exhausted",
         "EDI transaction error",
         "SSIS ETL job terminated unexpectedly after the connection pool limit was reached. Concurrent package executions exceeded the configured maximum. Pool sizing review needed."),
    (35, "P2", "In Progress", "ServiceBus queue backlog – dead-letter accumulation",
         "EDI transaction error",
         "Messages accumulating in the ServiceBus dead-letter queue. Consumer service is not processing messages. Lock expiry and retry exhaustion identified as root cause."),
    (36, "P3", "Open",        "Dotnet upgrade breaking change in EDI library",
         "EDI mapping issue",
         "Post dotnet upgrade, the EDI parsing library is throwing MethodNotFoundException. A deprecated API call in the serialization layer needs to be replaced with the new contract."),
    (37, "P2", "Pending",     "27x transaction mapping error on outbound file",
         "EDI mapping issue",
         "27x outbound file mapping is producing incorrect element values in the 2000B loop. Trading partner has rejected the last three transmissions citing format non-compliance."),
    (38, "P2", "Open",        "EUTF validation failed on inbound 837 file",
         "HIPAA validation failure",
         "EUTF validation layer rejected the inbound 837P file. Multiple ISA segments failed the EUTF rule set. File placed in error queue pending correction and resubmission."),
    (39, "P2", "In Progress", "835 remittance file not received from payer",
         "835 file not received",
         "Expected 835 remittance file from payer has not arrived within the agreed delivery window. Manual follow-up with trading partner initiated. Payment posting is on hold."),
    (40, "P2", "Open",        "834 enrollment file rejected by downstream processor",
         "HIPAA validation failure",
         "834 enrollment file failed HIPAA compliance validation at the downstream processor. Member demographic data contains invalid state code in the N4 segment."),
    (41, "P3", "Pending",     "Tumbleweed FTP transfer error – connection refused",
         "EDI partner connectivity",
         "Tumbleweed secure file transfer failed. Remote server refused connection on port 22. SSH key fingerprint has changed on partner server. Key update required in Tumbleweed config."),
    (42, "P3", "Open",        "PEX transaction rejected – prior auth mismatch",
         "EDI transaction error",
         "PEX prior authorization exchange transaction rejected by the receiving system. Authorization number referenced in the 278 request does not match any active prior auth record."),
    (43, "P2", "In Progress", "MEG batch processing error – file incomplete",
         "EDI transaction error",
         "MEG batch file arrived with truncated content. File size is 40% below the expected range. Partial records written to staging table causing downstream FK constraint violations."),
    (44, "P2", "Open",        "Tidal scheduler job missed trigger window",
         "EDI transaction error",
         "Critical Tidal batch job did not execute at the scheduled time. Trigger event was not received by the job server. Dependency chain has stalled 12 downstream processes."),
    (45, "P3", "Pending",     "MCG clinical criteria lookup returning null results",
         "EDI mapping issue",
         "MCG criteria engine is returning null for valid clinical codes. Cache invalidation event appears to have cleared the criteria lookup table. Cache warm-up required."),
    (46, "P2", "Open",        "ECRTP claim transformation step aborted",
         "EDI mapping issue",
         "ECRTP claim transformation pipeline aborted at step 3. Transformation rule set version mismatch between the orchestration engine and the rules repository."),
    (47, "P3", "In Progress", "Captiva document capture pipeline failure",
         "EDI transaction error",
         "Captiva OCR capture job failed to index incoming scanned documents. Document classification model returned low-confidence scores below the acceptance threshold."),
    (48, "P2", "Open",        "ACK service not responding to 997 functional ack",
         "Acknowledgement not received",
         "ACK service timed out while attempting to send a 997 functional acknowledgement. Service endpoint is reachable but processing queue is backlogged. Restart recommended."),
    (49, "P1", "In Progress", "837 claim file rejected at clearinghouse gateway",
         "837 claim rejection",
         "Clearinghouse gateway rejected the 837 professional claim file. Interchange control number is flagged as duplicate. Source system is generating non-unique ISA13 values."),
    (50, "P2", "Open",        "x12data parsing error on inbound transaction",
         "EDI mapping issue",
         "Inbound x12data file could not be parsed by the EDI engine. Segment terminator character in the ISA16 element differs from the actual delimiter used throughout the file."),
]

REPORTERS = [
    "Ravi Kumar", "Priya Sharma", "Anjali Singh", "Mohan Das",
    "Sunita Patel", "Vikram Nair", "Deepa Rao", "Kiran Reddy",
    "Amit Joshi", "Neha Gupta",
]

BASE_DATE = datetime(2026, 6, 2)

PRIORITY_COLOURS = {
    "P1": "FFC7CE",  # red
    "P2": "FFEB9C",  # amber
    "P3": "BDD7EE",  # blue
    "P4": "C6EFCE",  # green
}

STATUS_COLOURS = {
    "Open":        "FCE4D6",
    "In Progress": "E2EFDA",
    "Pending":     "FFF2CC",
}


# ── Excel generation ──────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # ── Column definitions: (header, width) ──────────────────────────────────
    columns = [
        ("Ticket_ID",     14),
        ("Date",          14),
        ("Reporter",      20),
        ("Status",        14),
        ("Priority",      10),
        ("Category",      32),
        ("Summary",       52),
        ("Description",   72),
        ("Assigned_Team", 16),
    ]

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="001529")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    body_font   = Font(name="Calibri", size=10)
    wrap_align  = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    random.seed(42)
    for row_num, (
        ticket_num, priority, status,
        summary, category, description
    ) in enumerate(TICKETS, start=2):

        ticket_id = f"TKT-{ticket_num:04d}"
        date_val  = (BASE_DATE + timedelta(days=(ticket_num - 1) % 25)).strftime("%Y-%m-%d")
        reporter  = REPORTERS[(ticket_num - 1) % len(REPORTERS)]

        row_data = [
            ticket_id, date_val, reporter,
            status, priority, category,
            summary, description, "EDI Team",
        ]

        prio_colour   = PatternFill("solid", fgColor=PRIORITY_COLOURS.get(priority, "FFFFFF"))
        status_colour = PatternFill("solid", fgColor=STATUS_COLOURS.get(status, "FFFFFF"))
        # Alternate row background for readability
        row_fill = PatternFill("solid", fgColor="F5F8FF" if row_num % 2 == 0 else "FFFFFF")

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font   = body_font
            cell.border = _thin_border()

            if col_idx in (5,):       # Priority column → colour-coded
                cell.fill      = prio_colour
                cell.alignment = center_align
            elif col_idx in (4,):     # Status column → colour-coded
                cell.fill      = status_colour
                cell.alignment = center_align
            elif col_idx in (1, 2, 9):  # ID / Date / Team → centred
                cell.alignment = center_align
                cell.fill      = row_fill
            else:
                cell.alignment = wrap_align
                cell.fill      = row_fill

        # Taller rows for description column
        ws.row_dimensions[row_num].height = 48

    # ── Freeze top row ────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # ── Legend sheet ─────────────────────────────────────────────────────────
    lg = wb.create_sheet("Legend")
    lg.column_dimensions["A"].width = 20
    lg.column_dimensions["B"].width = 50

    lg_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    lg.cell(1, 1, "Item").fill  = PatternFill("solid", fgColor="001529")
    lg.cell(1, 1).font          = lg_header_font
    lg.cell(1, 2, "Meaning").fill = PatternFill("solid", fgColor="001529")
    lg.cell(1, 2).font            = lg_header_font

    legend_rows = [
        ("P1 (Red)",    "Critical — immediate action required, system down or data at risk"),
        ("P2 (Amber)",  "High — significant business impact, same-day resolution expected"),
        ("P3 (Blue)",   "Medium — operational impact, resolution within 48 hours"),
        ("P4 (Green)",  "Low — cosmetic or minor, resolve within current sprint"),
        ("Open",        "Ticket received, not yet assigned to an engineer"),
        ("In Progress", "Engineer is actively working on the issue"),
        ("Pending",     "Awaiting information from reporter or trading partner"),
        ("EDI Team",    "All tickets in this file are owned by the EDI Team"),
    ]
    for ri, (item, meaning) in enumerate(legend_rows, start=2):
        ca = lg.cell(ri, 1, item)
        cb = lg.cell(ri, 2, meaning)
        for c in (ca, cb):
            c.font   = Font(name="Calibri", size=10)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        lg.row_dimensions[ri].height = 22

    # P1-P4 colour chips
    for ri, (prio, colour) in enumerate(PRIORITY_COLOURS.items(), start=2):
        if lg.cell(ri, 1).value and lg.cell(ri, 1).value.startswith(prio):
            lg.cell(ri, 1).fill = PatternFill("solid", fgColor=colour)

    wb.save(str(output_path))
    print(f"Saved → {output_path}  ({len(TICKETS)} tickets)")


if __name__ == "__main__":
    out = Path(__file__).resolve().parents[1] / "input" / "sample_tickets.xlsx"
    out.parent.mkdir(exist_ok=True)
    build_excel(out)
