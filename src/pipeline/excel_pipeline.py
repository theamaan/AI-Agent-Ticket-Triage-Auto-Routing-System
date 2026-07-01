"""
Excel Pipeline — reads input .xlsx, validates schema, returns list of ticket dicts.
Also writes enriched output .xlsx with AI prediction columns appended.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {"Summary", "Description"}

OUTPUT_COLUMNS = [
    "Ticket_ID", "Date", "Summary", "Description", "Reporter",
    # Original values and AI predictions side-by-side for easy comparison
    "Category", "AI_Category",
    "Priority", "AI_Priority",
    "Assigned_Team", "Assigned_Team_AI",
    "Status",
    "Assigned_To", "Assignee_Email",
    "ML_Confidence", "LLM_Confidence", "Final_Confidence",
    "Routing_Status", "Email_Sent",
]

# Human-readable column headers for the downloaded Excel
COLUMN_HEADERS = {
    "Category":        "Original_Category",
    "AI_Category":     "AI_Category",
    "Priority":        "Original_Priority",
    "AI_Priority":     "AI_Priority",
    "Assigned_Team":   "Original_Team",
    "Assigned_Team_AI":"AI_Assigned_Team",
}


def read_tickets(file_bytes: bytes, filename: str = "upload.xlsx") -> list[dict]:
    """Parse uploaded Excel bytes into a list of row dicts."""
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:
        raise ValueError(f"Cannot read Excel file '{filename}': {exc}") from exc

    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(
            f"Excel file is missing required columns: {missing}. "
            f"Found: {list(df.columns)}"
        )

    # Normalise column names
    df.columns = [str(c).strip() for c in df.columns]

    # Fill NaN with empty string for text fields
    for col in ["Summary", "Description", "Ticket_ID", "Reporter", "Date", "Status"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)

    # Generate Ticket_ID if missing.
    # Include a batch timestamp so each upload produces unique IDs and
    # never collides with tickets already stored in the database.
    if "Ticket_ID" not in df.columns or df["Ticket_ID"].str.strip().eq("").all():
        batch_ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        df["Ticket_ID"] = [f"TKT-{batch_ts}-{i+1:04d}" for i in range(len(df))]

    return df.to_dict("records")


def write_enriched_excel(results: list[dict]) -> bytes:
    """Serialise enriched ticket results to Excel bytes for download."""
    df = pd.DataFrame(results)

    # Reorder columns — only include those that exist in the data
    ordered = [c for c in OUTPUT_COLUMNS if c in df.columns]
    remaining = [c for c in df.columns if c not in ordered]
    df = df[ordered + remaining]

    # Rename columns to human-friendly headers (original vs AI side-by-side)
    df = df.rename(columns=COLUMN_HEADERS)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Triage_Results")

        # Apply conditional formatting via openpyxl
        ws = writer.sheets["Triage_Results"]
        _apply_styles(ws, df)

    buf.seek(0)
    return buf.read()


def _apply_styles(ws: Any, df: pd.DataFrame) -> None:
    """Colour-code priority cells (P1=red, P2=orange, P3=blue, P4=green)."""
    try:
        from openpyxl.styles import Font, PatternFill

        priority_colours = {
            "P1": "FFC7CE",  # red
            "P2": "FFEB9C",  # orange/yellow
            "P3": "BDD7EE",  # blue
            "P4": "C6EFCE",  # green
        }

        col_names = list(df.columns)
        priority_col_idx = None
        for i, name in enumerate(col_names):
            if name in ("AI_Priority", "Original_Priority", "Priority"):
                priority_col_idx = i + 1  # 1-based
                break

        if priority_col_idx is None:
            return

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            cell = ws.cell(row=row_idx, column=priority_col_idx)
            colour = priority_colours.get(str(cell.value), None)
            if colour:
                cell.fill = PatternFill(
                    start_color=colour, end_color=colour, fill_type="solid"
                )
    except Exception as exc:
        logger.debug("Could not apply Excel styles: %s", exc)
